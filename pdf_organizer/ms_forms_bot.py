from __future__ import annotations

import argparse
import csv
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from playwright.sync_api import Browser, Page


@dataclass(frozen=True)
class ColumnMapping:
    question: str
    type: str
    separator: str = ";"


@dataclass(frozen=True)
class BotConfig:
    form_url: str
    mapping: list[ColumnMapping]
    submit_label: str | None = None


def load_mapping(mapping_path: Path) -> list[ColumnMapping]:
    payload = json.loads(mapping_path.read_text(encoding="utf-8"))
    columns = payload.get("columns")
    if not isinstance(columns, list) or not columns:
        raise ValueError("O arquivo de mapeamento precisa conter 'columns' como lista não vazia.")

    mappings: list[ColumnMapping] = []
    for idx, item in enumerate(columns, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"Item {idx} de columns é inválido.")

        question = str(item.get("question", "")).strip()
        field_type = str(item.get("type", "")).strip().lower()
        separator = str(item.get("separator", ";"))

        if not question:
            raise ValueError(f"Item {idx} sem 'question'.")
        if field_type not in {"text", "radio", "checkbox", "dropdown"}:
            raise ValueError(
                f"Item {idx} com type inválido: {field_type}. Use text, radio, checkbox ou dropdown."
            )

        mappings.append(ColumnMapping(question=question, type=field_type, separator=separator))

    return mappings


def load_first_rows(
    spreadsheet_path: Path,
    limit: int = 60,
    expected_columns: int = 12,
    skip_header: bool = False,
) -> list[list[str]]:
    suffix = spreadsheet_path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(spreadsheet_path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(spreadsheet_path)
    else:
        raise ValueError("Formato de planilha não suportado. Use .csv ou .xlsx")

    normalized: list[list[str]] = []
    for row in rows:
        cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
        if not any(cleaned):
            continue
        if len(cleaned) < expected_columns:
            raise ValueError(
                f"Linha com {len(cleaned)} colunas encontrada, mas são esperadas {expected_columns}."
            )
        normalized.append(cleaned[:expected_columns])

    if skip_header and normalized:
        normalized = normalized[1:]

    normalized = normalized[:limit]

    if not normalized:
        raise ValueError("Nenhuma linha válida foi encontrada na planilha.")

    return normalized


def _read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.reader(fp)
        return [list(row) for row in reader]


def _read_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Para ler arquivos .xlsx, instale a dependência openpyxl.") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _find_question_scope(page: "Page", question: str):
    locator = page.locator(
        f"xpath=//*[contains(normalize-space(.), {json.dumps(question)})]"
        "/ancestor::*[@role='group' or @data-automation-id='questionItem' or contains(@class,'question')][1]"
    )
    if locator.count() == 0:
        raise RuntimeError(f"Não foi possível encontrar a pergunta '{question}'.")
    return locator.first


def _mark_choice(scope, value: str, multi: bool = False) -> None:
    parts = [x.strip() for x in value.split(";") if x.strip()] if multi else [value.strip()]
    for part in parts:
        if not part:
            continue
        option = scope.get_by_role("radio", name=re.compile(f"^{re.escape(part)}$", re.IGNORECASE))
        if option.count() > 0:
            option.first.check(force=True)
            continue

        checkbox = scope.get_by_role("checkbox", name=re.compile(f"^{re.escape(part)}$", re.IGNORECASE))
        if checkbox.count() > 0:
            checkbox.first.check(force=True)
            continue

        clickable = scope.get_by_text(re.compile(f"^{re.escape(part)}$", re.IGNORECASE))
        if clickable.count() == 0:
            raise RuntimeError(f"Opção '{part}' não encontrada.")
        clickable.first.click(force=True)


def _fill_field(page: "Page", mapping: ColumnMapping, value: str) -> None:
    scope = _find_question_scope(page, mapping.question)

    if mapping.type == "text":
        text_input = scope.locator("textarea, input[type='text'], input:not([type])")
        if text_input.count() == 0:
            raise RuntimeError(f"Campo de texto não encontrado para '{mapping.question}'.")
        text_input.first.fill(value)
        return

    if mapping.type == "radio":
        _mark_choice(scope, value, multi=False)
        return

    if mapping.type == "checkbox":
        parts = [x.strip() for x in value.split(mapping.separator) if x.strip()]
        _mark_choice(scope, ";".join(parts), multi=True)
        return

    if mapping.type == "dropdown":
        combo = scope.get_by_role("combobox")
        if combo.count() == 0:
            raise RuntimeError(f"Dropdown não encontrado para '{mapping.question}'.")
        combo.first.click(force=True)
        _mark_choice(page, value, multi=False)
        return

    raise RuntimeError(f"Tipo de campo não suportado: {mapping.type}")


def _submit_form(page: "Page", submit_label: str | None) -> None:
    if submit_label:
        button = page.get_by_role("button", name=re.compile(submit_label, re.IGNORECASE))
        if button.count() == 0:
            raise RuntimeError(f"Botão de envio '{submit_label}' não encontrado.")
        button.first.click(force=True)
        return

    button = page.get_by_role("button", name=re.compile(r"(enviar|submit|send|next)", re.IGNORECASE))
    if button.count() == 0:
        raise RuntimeError("Botão de envio não encontrado automaticamente.")
    button.first.click(force=True)


def detect_form_questions(page: "Page") -> list[str]:
    selectors = [
        "[data-automation-id='questionTitle']",
        "[data-automation-id='questionTitle'] *",
        "div[role='heading']",
        "h1, h2, h3, h4",
    ]

    found: list[str] = []
    seen: set[str] = set()
    for selector in selectors:
        nodes = page.locator(selector)
        for idx in range(nodes.count()):
            text = nodes.nth(idx).inner_text().strip()
            if not text or len(text) < 3:
                continue
            low = text.lower()
            if low in seen:
                continue
            if low in {"enviar", "submit", "next"}:
                continue
            seen.add(low)
            found.append(text)

    return found


def create_mapping_template(form_url: str, output_path: Path, headless: bool = False, limit: int = 12) -> None:
    from playwright.sync_api import sync_playwright

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=headless)
        try:
            page = browser.new_page()
            page.goto(form_url, wait_until="networkidle")
            page.wait_for_timeout(2000)
            questions = detect_form_questions(page)
        finally:
            browser.close()

    if not questions:
        raise RuntimeError("Não foi possível detectar perguntas automaticamente nesse formulário.")

    payload = {
        "columns": [
            {"question": question, "type": "text"}
            for question in questions[:limit]
        ]
    }
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def submit_rows(browser: "Browser", config: BotConfig, rows: list[list[str]], delay_ms: int = 700) -> None:
    for i, row in enumerate(rows, start=1):
        page = browser.new_page()
        page.goto(config.form_url, wait_until="networkidle")

        for col_idx, mapping in enumerate(config.mapping):
            _fill_field(page, mapping, row[col_idx])

        _submit_form(page, config.submit_label)
        time.sleep(delay_ms / 1000)
        page.close()
        print(f"[{i}/{len(rows)}] resposta enviada com sucesso")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="ms-form-bot",
        description="Preenche automaticamente um Microsoft Form usando as primeiras linhas de uma planilha.",
    )
    parser.add_argument("--form-url", required=True, help="URL de resposta do Microsoft Form")
    parser.add_argument("--sheet", help="Planilha CSV/XLSX com as respostas")
    parser.add_argument("--mapping", help="JSON com mapeamento das 12 colunas para perguntas")
    parser.add_argument("--limit", type=int, default=60, help="Quantidade de linhas a enviar (padrão: 60)")
    parser.add_argument("--delay-ms", type=int, default=700, help="Espera entre envios em milissegundos")
    parser.add_argument("--submit-label", default=None, help="Texto exato/parcial do botão de envio")
    parser.add_argument("--headless", action="store_true", help="Executa sem abrir janela do navegador")
    parser.add_argument(
        "--skip-header",
        action="store_true",
        help="Ignora a primeira linha da planilha (cabeçalho)",
    )
    parser.add_argument(
        "--init-mapping",
        help="Gera automaticamente um arquivo de mapeamento inicial com perguntas detectadas",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()

    if args.init_mapping:
        output = Path(args.init_mapping).expanduser().resolve()
        create_mapping_template(args.form_url, output, headless=args.headless)
        print(f"Mapeamento inicial criado em: {output}")
        return

    if not args.sheet or not args.mapping:
        raise SystemExit("Para enviar respostas, informe --sheet e --mapping.")

    sheet_path = Path(args.sheet).expanduser().resolve()
    mapping_path = Path(args.mapping).expanduser().resolve()

    if not sheet_path.exists():
        raise SystemExit(f"Planilha não encontrada: {sheet_path}")
    if not mapping_path.exists():
        raise SystemExit(f"Arquivo de mapeamento não encontrado: {mapping_path}")

    mapping = load_mapping(mapping_path)
    rows = load_first_rows(
        sheet_path,
        limit=args.limit,
        expected_columns=len(mapping),
        skip_header=args.skip_header,
    )

    config = BotConfig(
        form_url=args.form_url,
        mapping=mapping,
        submit_label=args.submit_label,
    )

    from playwright.sync_api import sync_playwright

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=args.headless)
        try:
            submit_rows(browser, config, rows, delay_ms=args.delay_ms)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
