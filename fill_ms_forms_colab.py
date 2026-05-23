#!/usr/bin/env python3
"""Preenche Microsoft Forms com dados de planilha Excel (Google Colab)."""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


DEFAULT_TIMEOUT_MS = 20_000
NEXT_LABELS = ["próximo", "proximo", "avançar", "avancar", "next"]


@dataclass
class FieldOption:
    label: str
    normalized: str
    locator_desc: str


@dataclass
class FormField:
    question_text: str
    normalized_question: str
    field_type: str
    required: bool
    container_selector: str
    input_selector: str | None = None
    options: list[FieldOption] = field(default_factory=list)


@dataclass
class MappingResult:
    excel_column: str
    form_question: str | None
    field_type: str | None
    status: str
    score: float
    reason: str


def normalize_text(value: str) -> str:
    value = value.replace("_x000D_", "\n")
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"\n+", "\n", value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().strip()
    value = re.sub(r"[^\w\s\n]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("_x000D_", "\n")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def split_multi_values(text: str) -> list[str]:
    parts = re.split(r"[;|\n]+", text)
    return [p.strip() for p in parts if p.strip()]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Preenche Microsoft Forms usando planilha .xlsx")
    parser.add_argument("--form-url", required=True)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--sheet-name")
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--dry-run", action="store_true", help="Padrão lógico se --submit não for usado")
    parser.add_argument("--submit", action="store_true", help="Habilita envio real")
    parser.add_argument("--headless", action="store_true")
    parser.add_argument("--screenshot-dir", default="./screenshots")
    parser.add_argument("--delay-ms", type=int, default=250)
    parser.add_argument("--submit-label", default="Enviar")
    parser.add_argument("--start-index", type=int, default=1)
    parser.add_argument("--mapping-file", default="form_mapping.json")
    parser.add_argument("--suggested-mapping-file", default="form_mapping_sugerido.json")
    parser.add_argument("--report-csv", default="relatorio_execucao.csv")
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    return parser.parse_args()


def read_spreadsheet(path: str, sheet_name: str | None, start_row: int) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [clean_cell_value(c.value) for c in ws[1]]

    used_col_idxs = [i for i, h in enumerate(headers) if h]
    filtered_headers = [headers[i] for i in used_col_idxs]
    rows: list[tuple[int, dict[str, str]]] = []

    for row_num in range(start_row, ws.max_row + 1):
        row_map: dict[str, str] = {}
        for col_idx, header in zip(used_col_idxs, filtered_headers):
            cell_val = clean_cell_value(ws.cell(row=row_num, column=col_idx + 1).value)
            if cell_val:
                row_map[header] = cell_val
        if row_map:
            rows.append((row_num, row_map))

    return filtered_headers, rows


def detect_login_or_block(page: Page) -> str | None:
    txt = normalize_text(page.locator("body").inner_text(timeout=3_000))
    if "sign in" in txt or "entrar" in txt and "microsoft" in txt:
        return "Formulário exige login/autenticação Microsoft."
    if "unusual traffic" in txt or "captcha" in txt:
        return "Possível bloqueio anti-automação/CAPTCHA detectado."
    return None


def collect_form_fields(page: Page) -> list[FormField]:
    script = """
() => {
  const questions = [];
  const candidates = Array.from(document.querySelectorAll('div[role="group"], fieldset, section'));
  let idx = 0;
  for (const c of candidates) {
    const questionEl = c.querySelector('span, legend, h1, h2, h3, label');
    if (!questionEl) continue;
    const qText = (questionEl.innerText || '').trim();
    if (!qText || qText.length < 2) continue;

    const textInput = c.querySelector('input[type="text"], input[type="email"], input[type="number"], textarea, input[type="date"]');
    const radios = Array.from(c.querySelectorAll('input[type="radio"]'));
    const checks = Array.from(c.querySelectorAll('input[type="checkbox"]'));
    const combo = c.querySelector('[role="combobox"], select');

    let type = 'unknown';
    if (textInput) {
      if (textInput.tagName.toLowerCase() === 'textarea') type = 'long_text';
      else if (textInput.getAttribute('type') === 'date') type = 'date';
      else type = 'short_text';
    } else if (radios.length > 0) {
      type = 'radio';
    } else if (checks.length > 0) {
      type = 'checkbox';
    } else if (combo) {
      type = 'dropdown';
    }

    const required = !!c.querySelector('[aria-label*="required" i], [aria-required="true"], .required');
    const cid = `autofield-${idx++}`;
    c.setAttribute('data-autofield-id', cid);

    const options = [];
    if (type === 'radio' || type === 'checkbox') {
      const nodes = type === 'radio' ? radios : checks;
      for (const n of nodes) {
        const id = n.id;
        let label = '';
        if (id) {
          const lbl = c.querySelector(`label[for="${id}"]`);
          if (lbl) label = (lbl.innerText || '').trim();
        }
        if (!label) {
          const nearest = n.closest('label, div[role="radio"], div[role="checkbox"]');
          if (nearest) label = (nearest.innerText || '').trim();
        }
        if (label) options.push({label});
      }
    }

    questions.push({
      question_text: qText,
      field_type: type,
      required,
      container_selector: `[data-autofield-id="${cid}"]`,
      options,
    });
  }
  return questions;
}
"""
    raw = page.evaluate(script)
    fields: list[FormField] = []
    for q in raw:
        fields.append(
            FormField(
                question_text=q["question_text"],
                normalized_question=normalize_text(q["question_text"]),
                field_type=q["field_type"],
                required=bool(q["required"]),
                container_selector=q["container_selector"],
                options=[FieldOption(label=o["label"], normalized=normalize_text(o["label"]), locator_desc=o["label"]) for o in q.get("options", [])],
            )
        )
    return fields


def build_mapping(headers: list[str], fields: list[FormField], manual_map: dict[str, str] | None) -> tuple[dict[str, FormField], list[MappingResult]]:
    mapped: dict[str, FormField] = {}
    report: list[MappingResult] = []

    norm_to_field = {f.normalized_question: f for f in fields}

    for h in headers:
        if manual_map and h in manual_map:
            target = manual_map[h]
            target_norm = normalize_text(target)
            field = norm_to_field.get(target_norm)
            if field:
                mapped[h] = field
                report.append(MappingResult(h, field.question_text, field.field_type, "mapped_manual", 1.0, "mapeamento manual"))
            else:
                report.append(MappingResult(h, None, None, "manual_not_found", 0.0, "pergunta manual não encontrada"))
            continue

        hn = normalize_text(h)
        exact = [f for f in fields if f.normalized_question == hn]
        if len(exact) == 1:
            mapped[h] = exact[0]
            report.append(MappingResult(h, exact[0].question_text, exact[0].field_type, "mapped_exact", 1.0, "match exato"))
            continue

        candidates = sorted(
            ((SequenceMatcher(None, hn, f.normalized_question).ratio(), f) for f in fields),
            key=lambda x: x[0],
            reverse=True,
        )
        top_score, top_field = candidates[0]
        second_score = candidates[1][0] if len(candidates) > 1 else 0.0

        if top_score >= 0.86 and (top_score - second_score) >= 0.08:
            mapped[h] = top_field
            report.append(MappingResult(h, top_field.question_text, top_field.field_type, "mapped_fuzzy", top_score, "match fuzzy confiável"))
        elif top_score >= 0.75:
            report.append(MappingResult(h, top_field.question_text, top_field.field_type, "ambiguous", top_score, "fuzzy ambíguo"))
        else:
            report.append(MappingResult(h, None, None, "not_found", top_score, "sem correspondência"))

    return mapped, report


def save_mapping_suggestion(path: str, rows: list[MappingResult]) -> None:
    payload = [r.__dict__ for r in rows]
    Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def append_report(report_path: str, row_number: int, status: str, msg: str, screenshot: str) -> None:
    exists = Path(report_path).exists()
    with open(report_path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if not exists:
            w.writerow(["numero_linha_planilha", "status", "mensagem", "screenshot", "timestamp"])
        w.writerow([row_number, status, msg, screenshot, datetime.now(timezone.utc).isoformat()])


def choose_single_option(options: list[FieldOption], wanted: str) -> FieldOption:
    nw = normalize_text(wanted)
    exact = [o for o in options if o.normalized == nw]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        raise ValueError(f"Opção ambígua: '{wanted}'")
    scored = sorted(((SequenceMatcher(None, nw, o.normalized).ratio(), o) for o in options), key=lambda x: x[0], reverse=True)
    if not scored or scored[0][0] < 0.86:
        raise ValueError(f"Opção não encontrada: '{wanted}'")
    if len(scored) > 1 and (scored[0][0] - scored[1][0]) < 0.08:
        raise ValueError(f"Opção ambígua: '{wanted}'")
    return scored[0][1]


def fill_field(page: Page, field: FormField, value: str, delay_ms: int) -> None:
    container = page.locator(field.container_selector)
    if field.field_type in {"short_text", "long_text", "date"}:
        inp = container.locator("input[type='text'], input[type='email'], input[type='number'], textarea, input[type='date']").first
        inp.fill(value)
    elif field.field_type == "radio":
        opt = choose_single_option(field.options, value)
        container.get_by_label(opt.label, exact=False).first.check()
    elif field.field_type == "checkbox":
        for part in split_multi_values(value):
            opt = choose_single_option(field.options, part)
            container.get_by_label(opt.label, exact=False).first.check()
    elif field.field_type == "dropdown":
        combo = container.locator("select").first
        if combo.count() > 0:
            combo.select_option(label=value)
        else:
            raise ValueError("Dropdown customizado não suportado sem seletor estável.")
    else:
        raise ValueError(f"Tipo de campo não suportado/detectado: {field.field_type} ({field.question_text})")

    if delay_ms > 0:
        page.wait_for_timeout(delay_ms)


def run() -> int:
    args = parse_args()
    dry_run = not args.submit
    Path(args.screenshot_dir).mkdir(parents=True, exist_ok=True)

    headers, rows = read_spreadsheet(args.xlsx, args.sheet_name, args.start_row)
    if args.start_index < 1:
        raise ValueError("--start-index deve ser >= 1")
    rows = rows[args.start_index - 1 :]
    if args.limit:
        rows = rows[: args.limit]

    if not rows:
        print("Nenhuma linha útil encontrada para processar.")
        return 0

    manual_map = None
    if Path(args.mapping_file).exists():
        manual_map = json.loads(Path(args.mapping_file).read_text(encoding="utf-8"))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        context = browser.new_context()

        # página inicial para detectar e mapear estrutura
        base_page = context.new_page()
        base_page.goto(args.form_url, wait_until="domcontentloaded", timeout=args.timeout_ms)
        block_reason = detect_login_or_block(base_page)
        if block_reason:
            print(f"ERRO: {block_reason}")
            return 2

        form_fields = collect_form_fields(base_page)
        print(f"Campos detectados: {len(form_fields)}")
        for f in form_fields:
            print(f"- {f.field_type} | {'OBRIGATÓRIO' if f.required else 'opc'} | {f.question_text}")

        mapping, mapping_rows = build_mapping(headers, form_fields, manual_map)
        save_mapping_suggestion(args.suggested_mapping_file, mapping_rows)

        bad_map = [r for r in mapping_rows if r.status in {"ambiguous", "not_found", "manual_not_found"}]
        if bad_map:
            print("Mapeamento incompleto/ambíguo. Verifique form_mapping_sugerido.json ou crie form_mapping.json.")
            for m in bad_map:
                print(f"- Coluna '{m.excel_column}': {m.status} ({m.reason})")
            return 3

        for idx, (excel_row_num, row_values) in enumerate(rows, start=1):
            print(f"\nProcessando linha da planilha {excel_row_num} (índice útil {idx})")
            page = context.new_page()
            screenshot_file = ""
            try:
                page.goto(args.form_url, wait_until="domcontentloaded", timeout=args.timeout_ms)
                block_reason = detect_login_or_block(page)
                if block_reason:
                    raise RuntimeError(block_reason)

                for col, val in row_values.items():
                    if col not in mapping:
                        continue
                    field = mapping[col]
                    print(f"  Preenchendo '{col}' -> '{field.question_text}' [{field.field_type}]")
                    fill_field(page, field, val, args.delay_ms)

                screenshot_file = str(Path(args.screenshot_dir) / f"linha_{excel_row_num}_antes_envio.png")
                page.screenshot(path=screenshot_file, full_page=True)

                submit_button = page.get_by_role("button", name=re.compile(rf"^{re.escape(args.submit_label)}$", re.IGNORECASE))
                next_button = page.get_by_role("button", name=re.compile("|".join(NEXT_LABELS), re.IGNORECASE))

                # Navegar por páginas, sem submeter em dry-run
                max_steps = 10
                for _ in range(max_steps):
                    if submit_button.count() > 0:
                        break
                    if next_button.count() > 0:
                        next_button.first.click()
                        page.wait_for_timeout(max(args.delay_ms, 300))
                        submit_button = page.get_by_role("button", name=re.compile(rf"^{re.escape(args.submit_label)}$", re.IGNORECASE))
                        next_button = page.get_by_role("button", name=re.compile("|".join(NEXT_LABELS), re.IGNORECASE))
                    else:
                        break

                if submit_button.count() == 0:
                    raise RuntimeError(f"Botão final '{args.submit_label}' não encontrado.")

                if dry_run:
                    end_shot = str(Path(args.screenshot_dir) / f"linha_{excel_row_num}_dry_run_final.png")
                    page.screenshot(path=end_shot, full_page=True)
                    print("  DRY-RUN: envio não realizado por segurança.")
                    append_report(args.report_csv, excel_row_num, "dry_run", "pré-validação concluída sem envio", end_shot)
                else:
                    submit_button.first.click()
                    page.wait_for_timeout(max(args.delay_ms, 500))
                    ok_shot = str(Path(args.screenshot_dir) / f"linha_{excel_row_num}_enviado.png")
                    page.screenshot(path=ok_shot, full_page=True)
                    print("  Submissão enviada com sucesso.")
                    append_report(args.report_csv, excel_row_num, "enviado", "submissão realizada", ok_shot)

            except (PlaywrightTimeoutError, Exception) as exc:
                err_shot = str(Path(args.screenshot_dir) / f"linha_{excel_row_num}_erro.png")
                try:
                    page.screenshot(path=err_shot, full_page=True)
                except Exception:
                    err_shot = ""
                msg = str(exc)
                print(f"  ERRO: {msg}")
                append_report(args.report_csv, excel_row_num, "erro", msg, err_shot)
            finally:
                page.close()

        browser.close()

    print("\nExecução concluída.")
    print(f"Relatório: {args.report_csv}")
    print(f"Mapeamento sugerido: {args.suggested_mapping_file}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
